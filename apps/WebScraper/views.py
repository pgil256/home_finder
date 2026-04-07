import json
import logging
from io import BytesIO
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db.models import Q
from apps.KeywordSelection.models import Keyword
from celery.result import AsyncResult
from celery import chain
from .models import PropertyListing
from .tasks.scrape_data import scrape_pinellas_properties, scrape_tax_data
from .tasks.sort_data import generate_sorted_properties
from .tasks.listings_pdf import generate_listing_pdf
from .tasks.visual_data import analyze_data
from .tasks.email_results import send_results_via_email

logger = logging.getLogger(__name__)

# Common data for views
PINELLAS_CITIES = [
    'Clearwater', 'St. Petersburg', 'Largo', 'Pinellas Park', 'Dunedin',
    'Palm Harbor', 'Tarpon Springs', 'Seminole', 'Safety Harbor', 'Oldsmar',
    'Gulfport', 'St. Pete Beach', 'Treasure Island', 'Madeira Beach',
    'Indian Rocks Beach', 'Belleair', 'Kenneth City', 'South Pasadena',
    'Indian Shores', 'Redington Beach'
]

PROPERTY_TYPES = [
    'Single Family', 'Condo', 'Townhouse', 'Multi-Family',
    'Vacant Land', 'Mobile Home', 'Commercial'
]


def build_processing_pipeline(search_criteria, limit=10, user_email=None):
    """Build and launch the complete Pinellas County property data pipeline.
    Returns the AsyncResult for the first task to enable progress tracking.
    """
    # Build the chain: Property data -> Tax data -> Reports
    task_chain = chain(
        scrape_pinellas_properties.s(search_criteria, limit),
        scrape_tax_data.s(),
        generate_sorted_properties.s(),
        generate_listing_pdf.s(),
        analyze_data.s(),
    )

    # Add email task if recipient provided
    if user_email:
        task_chain |= send_results_via_email.s(user_email)

    # Freeze the chain to assign task IDs before applying
    # This works in production (parent chain not available until execution)
    frozen_chain = task_chain.freeze()

    # Collect task IDs by traversing the frozen chain
    task_ids = []
    current = frozen_chain
    while current:
        task_ids.append(current.id)
        # For chains, the next task is in .parent (reverse order)
        current = getattr(current, 'parent', None)

    # Reverse so first task is at index 0
    task_ids.reverse()

    # Store chain info in cache BEFORE applying
    first_task_id = task_ids[0]
    cache.set(f'chain:{first_task_id}', {
        'task_ids': task_ids,
        'total_tasks': len(task_ids),
    }, timeout=3600)  # 1 hour TTL

    # Now apply the frozen chain
    task_chain.apply_async()

    return AsyncResult(first_task_id)


def web_scraper_view(request):
    """Main view for the property scraper interface"""
    if request.method == 'POST':
        # Get search criteria from the new unified search form
        property_types = request.POST.getlist('property_type')

        search_criteria = {
            'city': request.POST.get('city'),
            'zip_code': request.POST.get('zip_code'),
            'property_type': property_types[0] if len(property_types) == 1 else property_types if property_types else None,
            'min_value': request.POST.get('min_value'),
            'max_value': request.POST.get('max_value'),
            'bedrooms_min': request.POST.get('bedrooms_min'),
            'bathrooms_min': request.POST.get('bathrooms_min'),
            'year_built_after': request.POST.get('year_built_after'),
            'tax_status': request.POST.get('tax_status'),
            'sqft_min': request.POST.get('sqft_min'),
            'sqft_max': request.POST.get('sqft_max'),
        }

        # Remove empty values
        search_criteria = {k: v for k, v in search_criteria.items() if v}

        limit = int(request.POST.get('limit', 50))
        user_email = request.POST.get('email', '')

        # Start the pipeline - track the chain's first task for progress
        result = build_processing_pipeline(search_criteria, limit, user_email)

        return redirect('scraping-progress', task_id=result.id)

    # GET request - show the unified search interface
    context = {
        'cities': sorted(PINELLAS_CITIES),
        'property_types': PROPERTY_TYPES,
    }
    return render(request, 'WebScraper/search.html', context)


def scraping_progress(request, task_id):
    """View to track scraping progress"""
    context = {'task_id': task_id}
    return render(request, 'WebScraper/scraping-progress.html', context)


@csrf_exempt
def get_task_status(request, task_id):
    """API endpoint to get task status with chain progress aggregation."""
    # Check if this is part of a chain
    chain_info = cache.get(f'chain:{task_id}')

    if chain_info:
        return _get_chain_status(task_id, chain_info)

    # Single task progress (fallback)
    return _get_single_task_status(task_id)


def _get_single_task_status(task_id):
    """Get status for a single task."""
    task = AsyncResult(task_id)

    if task.state == 'PENDING':
        response = {
            'state': task.state,
            'current': 0,
            'total': 100,
            'status': 'Pending...'
        }
    elif task.state != 'FAILURE':
        info = task.info if isinstance(task.info, dict) else {}
        response = {
            'state': task.state,
            'current': info.get('current', 0),
            'total': info.get('total', 100),
            'status': info.get('status', '')
        }
        if task.state == 'SUCCESS':
            response['result'] = task.result
    else:
        response = {
            'state': task.state,
            'current': 100,
            'total': 100,
            'status': str(task.info),
        }

    return JsonResponse(response)


def _get_chain_status(first_task_id, chain_info):
    """Get aggregated status across all tasks in a chain."""
    task_ids = chain_info['task_ids']
    total_tasks = chain_info['total_tasks']

    # Find the currently active task and calculate overall progress
    completed_tasks = 0
    active_task = None
    active_task_index = 0
    last_result = None
    failed = False
    failure_info = None

    for i, tid in enumerate(task_ids):
        task = AsyncResult(tid)

        if task.state == 'FAILURE':
            failed = True
            failure_info = str(task.info)
            break
        elif task.state == 'SUCCESS':
            completed_tasks += 1
            last_result = task.result
        elif task.state in ('PENDING', 'STARTED', 'PROGRESS'):
            active_task = task
            active_task_index = i
            break

    if failed:
        return JsonResponse({
            'state': 'FAILURE',
            'current': 100,
            'total': 100,
            'status': failure_info,
        })

    # All tasks completed
    if completed_tasks == total_tasks:
        return JsonResponse({
            'state': 'SUCCESS',
            'current': 100,
            'total': 100,
            'status': 'All tasks completed',
            'result': last_result,
        })

    # Calculate overall progress
    # Each task contributes equally to total progress
    task_weight = 100 / total_tasks
    base_progress = completed_tasks * task_weight

    # Add progress from the active task
    task_progress = 0
    status = 'Processing...'

    if active_task:
        if active_task.state == 'PENDING':
            status = f'Waiting for task {active_task_index + 1}/{total_tasks}...'
        else:
            info = active_task.info if isinstance(active_task.info, dict) else {}
            task_progress = info.get('current', 0) / 100 * task_weight
            status = info.get('status', f'Running task {active_task_index + 1}/{total_tasks}...')

    overall_progress = int(base_progress + task_progress)

    return JsonResponse({
        'state': 'PROGRESS',
        'current': overall_progress,
        'total': 100,
        'status': status,
        'step': active_task_index + 1,
        'total_steps': total_tasks,
    })


def property_dashboard(request):
    """Property dashboard with filtering, sorting, and pagination"""
    # Start with all properties
    properties = PropertyListing.objects.all()

    # Get filter parameters
    city = request.GET.get('city')
    property_types_filter = request.GET.getlist('property_type')
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    beds = request.GET.get('beds')
    baths = request.GET.get('baths')
    year_built = request.GET.get('year_built')
    tax_status = request.GET.get('tax_status')
    sort = request.GET.get('sort', '-market_value')

    # Apply filters
    if city:
        properties = properties.filter(city__iexact=city)

    if property_types_filter:
        properties = properties.filter(property_type__in=property_types_filter)

    if min_price:
        try:
            properties = properties.filter(market_value__gte=float(min_price))
        except ValueError:
            pass

    if max_price:
        try:
            properties = properties.filter(market_value__lte=float(max_price))
        except ValueError:
            pass

    if beds and beds != '0':
        try:
            properties = properties.filter(bedrooms__gte=int(beds))
        except ValueError:
            pass

    if baths and baths != '0':
        try:
            properties = properties.filter(bathrooms__gte=float(baths))
        except ValueError:
            pass

    if year_built:
        try:
            properties = properties.filter(year_built__gte=int(year_built))
        except ValueError:
            pass

    if tax_status:
        properties = properties.filter(tax_status=tax_status)

    # Apply sorting
    valid_sort_fields = [
        'market_value', '-market_value',
        'created_at', '-created_at',
        'building_sqft', '-building_sqft',
        'year_built', '-year_built'
    ]
    if sort in valid_sort_fields:
        properties = properties.order_by(sort)
    else:
        properties = properties.order_by('-market_value')

    # Get total count before pagination
    total_count = properties.count()

    # Paginate
    paginator = Paginator(properties, 12)  # 12 properties per page
    page_number = request.GET.get('page', 1)
    properties_page = paginator.get_page(page_number)

    # Build search criteria summary
    search_criteria = {}
    if city:
        search_criteria['city'] = city
    if property_types_filter:
        search_criteria['property_types'] = property_types_filter

    context = {
        'properties': properties_page,
        'total_count': total_count,
        'cities': sorted(PINELLAS_CITIES),
        'property_types': PROPERTY_TYPES,
        'selected_property_types': property_types_filter,
        'search_criteria': search_criteria,
        'sort': sort,
    }
    return render(request, 'WebScraper/dashboard.html', context)


def property_detail(request, parcel_id):
    """Single property detail view"""
    property_obj = get_object_or_404(PropertyListing, parcel_id=parcel_id)

    # Get similar properties (same city, similar price range)
    similar_properties = PropertyListing.objects.filter(
        city=property_obj.city,
        property_type=property_obj.property_type
    ).exclude(parcel_id=parcel_id)

    if property_obj.market_value:
        min_price = float(property_obj.market_value) * 0.8
        max_price = float(property_obj.market_value) * 1.2
        similar_properties = similar_properties.filter(
            market_value__gte=min_price,
            market_value__lte=max_price
        )

    similar_properties = similar_properties[:4]

    context = {
        'property': property_obj,
        'similar_properties': similar_properties,
    }
    return render(request, 'WebScraper/property-detail.html', context)


def download_excel(request):
    """Generate and serve an Excel file of all properties on-demand."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    properties = PropertyListing.objects.all().order_by('-market_value')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Property Listings"

    headers = [
        'Parcel ID', 'Address', 'City', 'ZIP', 'Property Type',
        'Market Value', 'Assessed Value', 'Bedrooms', 'Bathrooms',
        'Sq Ft', 'Year Built', 'Lot Sq Ft', 'Tax Amount', 'Tax Status',
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for row_num, prop in enumerate(properties, 2):
        values = [
            prop.parcel_id, prop.address, prop.city, prop.zip_code,
            prop.property_type,
            float(prop.market_value) if prop.market_value else None,
            float(prop.assessed_value) if prop.assessed_value else None,
            prop.bedrooms, float(prop.bathrooms) if prop.bathrooms else None,
            prop.building_sqft, prop.year_built, prop.lot_sqft,
            float(prop.tax_amount) if prop.tax_amount else None,
            prop.tax_status,
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=row_num, column=col, value=value)

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="PropertyListings.xlsx"'
    return response


def download_pdf(request):
    """Generate PDF report with property listings and market analysis charts."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import pandas as pd
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
    )
    from reportlab.lib.units import inch

    properties = PropertyListing.objects.all().order_by('-market_value')[:200]

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        leftMargin=0.5*inch, rightMargin=0.5*inch,
        topMargin=0.5*inch, bottomMargin=0.5*inch,
    )

    story = []
    title_style = ParagraphStyle('Title', fontSize=20, alignment=TA_CENTER, spaceAfter=10)
    subtitle_style = ParagraphStyle('Subtitle', fontSize=11, alignment=TA_CENTER,
                                     textColor=rl_colors.grey, spaceAfter=30)

    story.append(Paragraph("Pinellas County Property Report", title_style))
    story.append(Paragraph(f"{properties.count()} properties &bull; Market Analysis", subtitle_style))

    # Property table
    header = ['Address', 'City', 'Type', 'Market Value', 'Beds', 'Baths', 'Sq Ft', 'Year', 'Tax']
    data = [header]
    for prop in properties:
        data.append([
            (prop.address or '')[:30],
            prop.city or '',
            (prop.property_type or '')[:20],
            f"${prop.market_value:,.0f}" if prop.market_value else 'N/A',
            str(prop.bedrooms or ''),
            str(prop.bathrooms or ''),
            f"{prop.building_sqft:,}" if prop.building_sqft else '',
            str(prop.year_built or ''),
            f"${prop.tax_amount:,.0f}" if prop.tax_amount else '',
        ])

    col_widths = [2*inch, 1.2*inch, 1.3*inch, 1.1*inch, 0.5*inch, 0.5*inch, 0.8*inch, 0.6*inch, 0.9*inch]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor("#1e3a5f")),
        ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#e2e8f0")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#f7fafc")]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    # Build DataFrame for charts (convert Decimal fields to float)
    raw = list(properties.values(
        'market_value', 'assessed_value', 'building_sqft', 'property_type',
        'year_built', 'tax_amount', 'city', 'bedrooms', 'bathrooms'
    ))
    df = pd.DataFrame(raw)
    for col in ['market_value', 'assessed_value', 'tax_amount', 'bathrooms']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    chart_colors = ['#1e3a5f', '#2c5282', '#38a169', '#dd6b20', '#c53030',
                    '#805ad5', '#d69e2e', '#319795', '#e53e3e', '#3182ce']

    def add_chart(fig, story):
        """Save matplotlib figure to BytesIO and add to ReportLab story."""
        img_buf = BytesIO()
        fig.savefig(img_buf, format='png', dpi=120, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        img_buf.seek(0)
        story.append(PageBreak())
        story.append(Image(img_buf, width=9*inch, height=5.5*inch))

    # Chart 1: Distribution of Market Values
    prices = df['market_value'].dropna()
    if len(prices) > 0:
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.hist(prices, bins=20, color=chart_colors[0], edgecolor='white', alpha=0.85)
        mean_price = prices.mean()
        ax.axvline(mean_price, color=chart_colors[2], linestyle='--', linewidth=2,
                   label=f'Mean: ${mean_price:,.0f}')
        ax.set_title('Distribution of Market Values', fontsize=14, fontweight='bold', pad=15)
        ax.set_xlabel('Market Value')
        ax.set_ylabel('Number of Properties')
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'${x:,.0f}'))
        ax.legend()
        plt.tight_layout()
        add_chart(fig, story)

    # Chart 2: Home Size vs Market Value
    plot_df = df[['building_sqft', 'market_value']].dropna()
    if len(plot_df) > 2:
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.scatter(plot_df['building_sqft'], plot_df['market_value'],
                   c=chart_colors[1], alpha=0.6, s=80, edgecolors='white')
        import numpy as np
        z = np.polyfit(plot_df['building_sqft'], plot_df['market_value'], 1)
        p = np.poly1d(z)
        x_line = np.linspace(plot_df['building_sqft'].min(), plot_df['building_sqft'].max(), 100)
        ax.plot(x_line, p(x_line), color=chart_colors[2], linestyle='--', linewidth=2, label='Trend')
        ax.set_title('Home Size vs. Market Value', fontsize=14, fontweight='bold', pad=15)
        ax.set_xlabel('Square Footage')
        ax.set_ylabel('Market Value')
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'${x:,.0f}'))
        ax.legend()
        plt.tight_layout()
        add_chart(fig, story)

    # Chart 3: Property Types
    type_counts = df['property_type'].value_counts().head(10)
    if len(type_counts) > 0:
        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(range(len(type_counts)), type_counts.values,
                      color=[chart_colors[i % len(chart_colors)] for i in range(len(type_counts))],
                      edgecolor='white')
        ax.set_title('Listings by Property Type', fontsize=14, fontweight='bold', pad=15)
        ax.set_xticks(range(len(type_counts)))
        ax.set_xticklabels(type_counts.index, rotation=45, ha='right', fontsize=8)
        ax.set_ylabel('Count')
        for bar in bars:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3,
                    str(int(bar.get_height())), ha='center', va='bottom', fontsize=9)
        plt.tight_layout()
        add_chart(fig, story)

    # Chart 4: Year Built Distribution
    years = df['year_built'].dropna()
    if len(years) > 0:
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.hist(years, bins=15, color=chart_colors[3], edgecolor='white', alpha=0.85)
        ax.set_title('Distribution of Year Built', fontsize=14, fontweight='bold', pad=15)
        ax.set_xlabel('Year Built')
        ax.set_ylabel('Number of Properties')
        plt.tight_layout()
        add_chart(fig, story)

    # Chart 5: Market Value vs Assessed Value
    val_df = df[['market_value', 'assessed_value']].dropna()
    val_df = val_df[(val_df['market_value'] > 0) & (val_df['assessed_value'] > 0)]
    if len(val_df) > 0:
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.scatter(val_df['assessed_value'], val_df['market_value'],
                   c=chart_colors[1], alpha=0.6, s=80, edgecolors='white')
        max_val = max(val_df['market_value'].max(), val_df['assessed_value'].max())
        min_val = min(val_df['market_value'].min(), val_df['assessed_value'].min())
        ax.plot([min_val, max_val], [min_val, max_val], color='grey',
                linestyle='--', linewidth=1.5, label='Equal Value', alpha=0.7)
        ax.set_title('Market Value vs. Assessed Value', fontsize=14, fontweight='bold', pad=15)
        ax.set_xlabel('Assessed Value')
        ax.set_ylabel('Market Value')
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'${x:,.0f}'))
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'${x:,.0f}'))
        ax.legend()
        plt.tight_layout()
        add_chart(fig, story)

    # Chart 6: Average Price by City
    city_stats = df.groupby('city')['market_value'].agg(['mean', 'count']).dropna()
    city_stats = city_stats[city_stats['count'] >= 1].sort_values('mean', ascending=False).head(10)
    if len(city_stats) > 0:
        fig, ax = plt.subplots(figsize=(10, 6))
        bars = ax.bar(range(len(city_stats)), city_stats['mean'].values,
                      color=[chart_colors[i % len(chart_colors)] for i in range(len(city_stats))],
                      edgecolor='white')
        ax.set_title('Average Market Value by City', fontsize=14, fontweight='bold', pad=15)
        ax.set_xticks(range(len(city_stats)))
        ax.set_xticklabels(city_stats.index, rotation=45, ha='right')
        ax.set_ylabel('Average Market Value')
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'${x:,.0f}'))
        for i, bar in enumerate(bars):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height(),
                    f'n={int(city_stats["count"].iloc[i])}', ha='center', va='bottom', fontsize=8)
        plt.tight_layout()
        add_chart(fig, story)

    # Chart 7: Tax Burden Analysis
    tax_df = df[['market_value', 'tax_amount']].dropna()
    tax_df = tax_df[(tax_df['market_value'] > 0) & (tax_df['tax_amount'] > 0)]
    if len(tax_df) > 0:
        tax_df['tax_rate'] = (tax_df['tax_amount'] / tax_df['market_value']) * 100
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.scatter(tax_df['market_value'], tax_df['tax_rate'],
                   c=chart_colors[3], alpha=0.6, s=80, edgecolors='white')
        avg_rate = tax_df['tax_rate'].mean()
        ax.axhline(avg_rate, color=chart_colors[4], linestyle='--', linewidth=2,
                   label=f'Avg: {avg_rate:.2f}%')
        ax.axhline(1.8, color=chart_colors[2], linestyle=':', linewidth=1.5,
                   label='FL Avg: 1.8%', alpha=0.7)
        ax.set_title('Tax Burden Analysis', fontsize=14, fontweight='bold', pad=15)
        ax.set_xlabel('Market Value')
        ax.set_ylabel('Effective Tax Rate (%)')
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f'${x:,.0f}'))
        ax.legend()
        plt.tight_layout()
        add_chart(fig, story)

    doc.build(story)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="PropertyReport.pdf"'
    return response