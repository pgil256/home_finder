import logging
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from celery import shared_task
from django.conf import settings
from apps.KeywordSelection.models import Keyword
from apps.WebScraper.models import PropertyListing
from celery_progress.backend import ProgressRecorder


# Setup logging
logger = logging.getLogger(__name__)

# Reports directory in media
REPORTS_DIR = os.path.join(settings.MEDIA_ROOT, 'reports')
os.makedirs(REPORTS_DIR, exist_ok=True)
logging.basicConfig(
    level=logging.DEBUG, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)


def fetch_property_listings():
    logger.debug("Fetching property listings based on priority")
    keywords = Keyword.objects.exclude(priority=0).order_by("-priority")
    priority_columns = [(keyword.listing_field, keyword.name) for keyword in keywords]
    priority_fields = [field for field, _ in priority_columns]

    all_fields = {field.name for field in PropertyListing._meta.get_fields()}
    non_priority_fields = sorted(all_fields - set(priority_fields))

    fields = priority_fields + non_priority_fields
    if None in fields:
        fields = [
            field for field in fields if field is not None
        ]  # Safeguard against None fields

    listings = PropertyListing.objects.values_list(*fields, named=True)
    logger.info(f"Database contains {len(listings)} property listings")

    # Log sample data for debugging
    if listings:
        sample = listings[0]
        logger.info(f"Sample listing: parcel_id={getattr(sample, 'parcel_id', 'N/A')}, "
                    f"address={getattr(sample, 'address', 'N/A')}, "
                    f"market_value={getattr(sample, 'market_value', 'N/A')}")

    return [(field, field) for field in fields], listings


def strip_timezone(value):
    """Strip timezone info from datetime objects for Excel compatibility."""
    if hasattr(value, 'tzinfo') and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def generate_spreadsheet(columns, listings):
    logger.info("Generating spreadsheet for property listings")
    df = pd.DataFrame(list(listings), columns=[name for _, name in columns])

    # Strip timezone info from datetime columns (Excel doesn't support timezones)
    for col in df.columns:
        if df[col].dtype == 'datetime64[ns, UTC]' or df[col].apply(
            lambda x: hasattr(x, 'tzinfo') and x is not None and getattr(x, 'tzinfo', None) is not None
        ).any():
            df[col] = df[col].apply(strip_timezone)
    wb = Workbook()
    ws = wb.active
    header_font = Font(bold=True, color="FFFFFF")
    cell_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )

    for col_num, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = cell_border
        cell.fill = header_fill

    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)

    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = max_length

    filename = "PropertyListings.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb.save(filepath)
    logger.info(f"Spreadsheet saved to {filepath}")

    # Verify file was created
    if os.path.exists(filepath):
        file_size = os.path.getsize(filepath)
        logger.info(f"File verified: {filepath} ({file_size} bytes)")
    else:
        logger.error(f"File NOT created: {filepath}")

    return filepath


def quick_sort(arr, compare_func):
    if len(arr) <= 1:
        return arr
    pivot = arr[0]
    less = [x for x in arr[1:] if compare_func(x, pivot) < 0]
    greater = [x for x in arr[1:] if compare_func(x, pivot) >= 0]
    return quick_sort(less, compare_func) + [pivot] + quick_sort(greater, compare_func)


def compare_properties(x, y, user_preferences, priority_fields):
    score_x, score_y = 0, 0

    def compare_range_or_max(x_attr, y_attr, pref_min, pref_max, is_max=False):
        nonlocal score_x, score_y
        value_min = float(pref_min) if pref_min else float("-inf")
        value_max = float(pref_max) if pref_max else float("inf")
        if not is_max:
            if value_min <= getattr(x, x_attr) <= value_max:
                score_x += 1
            if value_min <= getattr(y, y_attr) <= value_max:
                score_y += 1
        else:
            if getattr(x, x_attr) <= value_max:
                score_x += 1
            if getattr(y, y_attr) <= value_max:
                score_y += 1

    def compare_exact(x_attr, y_attr, preference):
        nonlocal score_x, score_y
        if getattr(x, x_attr) == preference:
            score_x += 1
        if getattr(y, y_attr) == preference:
            score_y += 1

    for field in priority_fields:
        pref_min = user_preferences.get(f"{field}_min", None)
        pref_max = user_preferences.get(f"{field}_max", None)
        is_max = "+" in str(pref_max) if pref_max else False

        if pref_min or pref_max:
            compare_range_or_max(field, field, pref_min, pref_max, is_max)
        elif user_preferences.get(field):
            compare_exact(field, field, user_preferences[field])

    return score_x - score_y


@shared_task(bind=True)
def generate_sorted_properties(self, tax_result):
    progress_recorder = ProgressRecorder(self)

    # Extract search_criteria and limit from the chain result
    scrape_config = tax_result.get('search_criteria', {}) if isinstance(tax_result, dict) else {}
    limit = tax_result.get('limit', 10) if isinstance(tax_result, dict) else 10

    if not scrape_config:
        logger.warning("No scrape configuration provided, using empty config")
        scrape_config = {}

    columns, listings = fetch_property_listings()
    progress_recorder.set_progress(25, 100, description="Fetched property listings")

    priority_fields = [col[0] for col in columns if col[0] != "link"]
    logger.debug(f"Sorting properties based on user preferences: {scrape_config}")
    sorted_properties = quick_sort(
        listings, lambda x, y: compare_properties(x, y, scrape_config, priority_fields)
    )
    progress_recorder.set_progress(35, 100, description="Sorted properties")

    logger.info("Sorted properties, generating spreadsheet")
    excel_filepath = generate_spreadsheet(columns, sorted_properties)
    progress_recorder.set_progress(45, 100, description="Generated spreadsheet")

    logger.info("Generating PDF for sorted properties")
    # Code to generate PDF goes here
    progress_recorder.set_progress(50, 100, description="Generated PDF")

    logger.info(f"Top {limit} sorted properties generated")
    return {
        'sorted_properties': sorted_properties[:limit],
        'columns': columns,
        'excel_path': excel_filepath
    }
