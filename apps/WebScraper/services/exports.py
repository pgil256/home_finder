"""Excel + PDF market analysis exports.

Both honor the dashboard's active filters so the export reflects exactly
the slice of the market the user is looking at. Intended audience is a
prospective buyer or investor, not someone wanting a raw data dump.
"""

from __future__ import annotations

import logging
from datetime import datetime
from io import BytesIO
from typing import Any

from django.db.models import Avg, Count, Max, Min, Sum
from django.http import HttpResponse

from ..models import PropertyListing
from .filtering import apply_filters

logger = logging.getLogger(__name__)

# Caps so neither export OOMs the Vercel function (60s budget):
MAX_EXCEL_PROPERTIES = 5000
MAX_PDF_PROPERTIES = 200

# bedrooms/bathrooms intentionally excluded — PCPAO doesn't publish them.
LISTING_COLUMNS = (
    ('parcel_id', 'Parcel ID', None),
    ('address', 'Address', None),
    ('city', 'City', None),
    ('zip_code', 'ZIP', None),
    ('property_type', 'Property Type', None),
    ('market_value', 'Market Value', '"$"#,##0'),
    ('assessed_value', 'Assessed Value', '"$"#,##0'),
    ('building_sqft', 'Sq Ft', '#,##0'),
    ('year_built', 'Year Built', '0'),
    ('lot_sqft', 'Lot Sq Ft', '#,##0'),
    ('tax_amount', 'Tax Amount', '"$"#,##0'),
    ('tax_status', 'Tax Status', None),
)


def _filtered_queryset(request):
    if request is None:
        return PropertyListing.objects.all()
    qs, _, _ = apply_filters(request)
    return qs


def _summarize_filters(request) -> list[tuple[str, str]]:
    """Return a list of (label, value) describing what filters are active."""
    if request is None:
        return [('Scope', 'All Pinellas County properties')]
    out: list[tuple[str, str]] = []
    g = request.GET
    if g.get('q'):
        out.append(('Keyword', g['q']))
    if g.get('city'):
        out.append(('City', g['city']))
    if g.get('zip_code'):
        out.append(('ZIP', g['zip_code']))
    types = g.getlist('property_type')
    if types:
        out.append(('Property type', ', '.join(types)))
    if g.get('min_price') or g.get('max_price'):
        lo = f'${int(g["min_price"]):,}' if g.get('min_price') else 'any'
        hi = f'${int(g["max_price"]):,}' if g.get('max_price') else 'any'
        out.append(('Price range', f'{lo} – {hi}'))
    if g.get('year_built'):
        out.append(('Built after', g['year_built']))
    if g.get('min_sqft') or g.get('max_sqft'):
        lo = f'{int(g["min_sqft"]):,}' if g.get('min_sqft') else 'any'
        hi = f'{int(g["max_sqft"]):,}' if g.get('max_sqft') else 'any'
        out.append(('Square footage', f'{lo} – {hi}'))
    if g.get('min_lot_sqft') or g.get('max_lot_sqft'):
        lo = f'{int(g["min_lot_sqft"]):,}' if g.get('min_lot_sqft') else 'any'
        hi = f'{int(g["max_lot_sqft"]):,}' if g.get('max_lot_sqft') else 'any'
        out.append(('Lot square footage', f'{lo} – {hi}'))
    if g.get('min_tax_amount') or g.get('max_tax_amount'):
        lo = f'${int(g["min_tax_amount"]):,}' if g.get('min_tax_amount') else 'any'
        hi = f'${int(g["max_tax_amount"]):,}' if g.get('max_tax_amount') else 'any'
        out.append(('Annual tax amount', f'{lo} – {hi}'))
    if not out:
        out.append(('Scope', 'All Pinellas County properties'))
    return out


def _market_stats(qs) -> dict[str, Any]:
    """Compute headline market stats over the (already filtered) queryset."""
    aggs = qs.aggregate(
        n=Count('id'),
        mean_value=Avg('market_value'),
        min_value=Min('market_value'),
        max_value=Max('market_value'),
        sum_value=Sum('market_value'),
        mean_sqft=Avg('building_sqft'),
        mean_year=Avg('year_built'),
    )
    # Median needs a separate query (no DB-agnostic aggregate in Django).
    n_with_value = qs.exclude(market_value__isnull=True).count()
    median = None
    if n_with_value:
        offset = n_with_value // 2
        median_qs = (
            qs.exclude(market_value__isnull=True)
            .order_by('market_value')
            .values_list('market_value', flat=True)[offset : offset + 1]
        )
        median = next(iter(median_qs), None)
    aggs['median_value'] = median
    return aggs


def _by_property_type(qs, limit: int = 15):
    """Return per-property-type counts and price stats, top N by count."""
    return list(
        qs.values('property_type')
        .annotate(
            count=Count('id'),
            mean_value=Avg('market_value'),
            mean_sqft=Avg('building_sqft'),
        )
        .order_by('-count')[:limit]
    )


def _by_city(qs, limit: int = 25):
    return list(
        qs.values('city')
        .annotate(
            count=Count('id'),
            mean_value=Avg('market_value'),
        )
        .order_by('-count')[:limit]
    )


# -------- Excel ---------------------------------------------------------


def generate_excel_response(request=None) -> HttpResponse:
    """3-sheet workbook: Summary, Listings, By Property Type (+ By City if multi-city)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    qs_full = _filtered_queryset(request)
    listings = qs_full.only(*[c[0] for c in LISTING_COLUMNS]).order_by('-market_value')[:MAX_EXCEL_PROPERTIES]
    stats = _market_stats(qs_full)
    by_type = _by_property_type(qs_full)
    cities = _by_city(qs_full)
    multi_city = sum(1 for c in cities if c['city']) > 1
    filter_summary = _summarize_filters(request)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1E3A5F')
    title_font = Font(bold=True, size=14)

    # ---- Sheet 1: Summary
    ws = wb.active
    ws.title = 'Summary'
    ws['A1'] = 'Pinellas Property Market Snapshot'
    ws['A1'].font = title_font
    ws['A2'] = f'Generated {datetime.utcnow():%Y-%m-%d %H:%M UTC}'
    ws['A2'].font = Font(italic=True, color='808080')

    row = 4
    ws.cell(row, 1, 'Filters applied').font = Font(bold=True)
    row += 1
    for label, value in filter_summary:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        row += 1

    row += 1
    ws.cell(row, 1, 'Headline numbers').font = Font(bold=True)
    row += 1
    rows = [
        ('Properties matched', f'{stats["n"]:,}'),
        ('Median market value', _money(stats['median_value'])),
        ('Mean market value', _money(stats['mean_value'])),
        ('Min market value', _money(stats['min_value'])),
        ('Max market value', _money(stats['max_value'])),
        ('Total market value', _money(stats['sum_value'])),
        ('Mean square footage', _int(stats['mean_sqft'])),
        ('Mean year built', _int(stats['mean_year'])),
    ]
    for label, value in rows:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value).alignment = Alignment(horizontal='right')
        row += 1

    if listings.count() < stats['n']:
        row += 1
        cell = ws.cell(
            row,
            1,
            f'Note: the Listings sheet shows the top {MAX_EXCEL_PROPERTIES:,} of '
            f'{stats["n"]:,} matching properties (sorted by value, highest first).',
        )
        cell.font = Font(italic=True, color='808080')

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 36

    # ---- Sheet 2: Listings
    ws = wb.create_sheet('Listings')
    for col_idx, (_, header, _) in enumerate(LISTING_COLUMNS, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'

    for row_idx, prop in enumerate(listings, 2):
        for col_idx, (attr, _, fmt) in enumerate(LISTING_COLUMNS, 1):
            value = getattr(prop, attr, None)
            if value is not None and attr in ('market_value', 'assessed_value', 'tax_amount'):
                value = float(value)
            cell = ws.cell(row_idx, col_idx, value)
            if fmt:
                cell.number_format = fmt

    for col_idx, (_, header, _) in enumerate(LISTING_COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        # Approximate width: longest of header / 12 chars min / 40 max
        ws.column_dimensions[col_letter].width = max(12, min(len(header) + 4, 40))

    # ---- Sheet 3: By Property Type
    ws = wb.create_sheet('By Property Type')
    headers = ['Property Type', 'Count', 'Mean Value', 'Mean Sq Ft']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    for row_idx, row_data in enumerate(by_type, 2):
        ws.cell(row_idx, 1, row_data['property_type'] or 'Unknown')
        ws.cell(row_idx, 2, row_data['count']).number_format = '#,##0'
        ws.cell(row_idx, 3, float(row_data['mean_value'] or 0)).number_format = '"$"#,##0'
        ws.cell(row_idx, 4, int(row_data['mean_sqft'] or 0)).number_format = '#,##0'
    ws.column_dimensions['A'].width = 40
    for c in 'BCD':
        ws.column_dimensions[c].width = 16

    # ---- Sheet 4: By City (only when more than one city in scope)
    if multi_city:
        ws = wb.create_sheet('By City')
        headers = ['City', 'Count', 'Mean Value']
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'
        for row_idx, row_data in enumerate(cities, 2):
            ws.cell(row_idx, 1, row_data['city'] or 'Unknown')
            ws.cell(row_idx, 2, row_data['count']).number_format = '#,##0'
            ws.cell(row_idx, 3, float(row_data['mean_value'] or 0)).number_format = '"$"#,##0'
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'PinellasMarket_{datetime.utcnow():%Y%m%d}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# -------- PDF -----------------------------------------------------------


def generate_pdf_response(request=None) -> HttpResponse:
    """PDF report: filter context, headline stats, top listings, market charts."""
    import matplotlib

    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import numpy as np
    import pandas as pd
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        Image,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    qs_full = _filtered_queryset(request)
    listings = qs_full.only(
        'parcel_id',
        'address',
        'city',
        'property_type',
        'market_value',
        'assessed_value',
        'building_sqft',
        'year_built',
        'tax_amount',
    ).order_by('-market_value')[:MAX_PDF_PROPERTIES]
    stats = _market_stats(qs_full)
    filter_summary = _summarize_filters(request)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(letter),
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )
    story: list[Any] = []

    title_style = ParagraphStyle('Title', fontSize=22, alignment=TA_CENTER, spaceAfter=4)
    subtitle_style = ParagraphStyle(
        'Subtitle',
        fontSize=11,
        alignment=TA_CENTER,
        textColor=rl_colors.grey,
        spaceAfter=18,
    )
    h_style = ParagraphStyle('H', fontSize=14, spaceBefore=8, spaceAfter=6)
    body_style = ParagraphStyle('Body', fontSize=10, alignment=TA_LEFT, spaceAfter=4)

    story.append(Paragraph('Pinellas Property Market Report', title_style))
    story.append(
        Paragraph(
            f'{stats["n"]:,} matching properties &bull; {datetime.utcnow():%B %d, %Y}',
            subtitle_style,
        )
    )

    # Filter context + headline stats panel (two-column table)
    summary_rows = [
        [Paragraph('<b>Filters applied</b>', body_style), Paragraph('<b>Headline numbers</b>', body_style)],
    ]
    fmts = [
        ('Properties matched', f'{stats["n"]:,}'),
        ('Median market value', _money(stats['median_value'])),
        ('Mean market value', _money(stats['mean_value'])),
        ('Min / Max', f'{_money(stats["min_value"])} / {_money(stats["max_value"])}'),
        ('Mean square footage', _int(stats['mean_sqft'])),
    ]
    rows_count = max(len(filter_summary), len(fmts))
    for i in range(rows_count):
        left = ''
        if i < len(filter_summary):
            label, value = filter_summary[i]
            left = f'<b>{label}:</b> {value}'
        right = ''
        if i < len(fmts):
            label, value = fmts[i]
            right = f'<b>{label}:</b> {value}'
        summary_rows.append([Paragraph(left, body_style), Paragraph(right, body_style)])

    summary_table = Table(summary_rows, colWidths=[5 * inch, 5 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#f1f5f9')),
                ('BOX', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#cbd5e1')),
                ('INNERGRID', (0, 0), (-1, -1), 0.25, rl_colors.HexColor('#e2e8f0')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 0.25 * inch))

    # Top listings table (bedrooms/bathrooms columns dropped — always empty)
    story.append(
        Paragraph(
            f'Top {min(len(listings), MAX_PDF_PROPERTIES)} properties by market value',
            h_style,
        )
    )
    header = ['Address', 'City', 'Type', 'Market Value', 'Sq Ft', 'Year', 'Annual Tax']
    data = [header]
    for prop in listings:
        data.append(
            [
                (prop.address or '')[:32],
                prop.city or '',
                (prop.property_type or '')[:24],
                _money(prop.market_value),
                f'{prop.building_sqft:,}' if prop.building_sqft else '',
                str(prop.year_built or ''),
                _money(prop.tax_amount),
            ]
        )
    table = Table(
        data,
        colWidths=[2.3 * inch, 1.3 * inch, 1.7 * inch, 1.1 * inch, 0.7 * inch, 0.6 * inch, 1.0 * inch],
        repeatRows=1,
    )
    table.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#1e3a5f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#e2e8f0')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor('#f7fafc')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)

    # ---- Charts page
    chart_colors = [
        '#1e3a5f',
        '#2c5282',
        '#38a169',
        '#dd6b20',
        '#c53030',
        '#805ad5',
        '#d69e2e',
        '#319795',
        '#e53e3e',
        '#3182ce',
    ]

    raw = list(
        qs_full.values(
            'market_value',
            'assessed_value',
            'building_sqft',
            'property_type',
            'year_built',
            'tax_amount',
            'city',
        )[:50_000]
    )
    df = pd.DataFrame(raw)
    for col in ('market_value', 'assessed_value', 'tax_amount'):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    def _add_chart(fig: Any, title: str = ''):
        img_buf = BytesIO()
        fig.savefig(img_buf, format='png', dpi=110, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        img_buf.seek(0)
        story.append(PageBreak())
        if title:
            story.append(Paragraph(title, h_style))
        story.append(Image(img_buf, width=9.5 * inch, height=5.0 * inch))

    # Chart 1: Distribution of Market Values (capped at 99th percentile so
    # one $300M hospital doesn't squash the rest of the histogram)
    prices = df['market_value'].dropna()
    prices = prices[prices > 0]
    if len(prices):
        cap = prices.quantile(0.99)
        clean = prices[prices <= cap]
        fig, ax = plt.subplots(figsize=(11, 5.5))
        ax.hist(clean, bins=40, color=chart_colors[0], edgecolor='white', alpha=0.85)
        med = prices.median()
        mean = prices.mean()
        ax.axvline(med, color=chart_colors[2], linestyle='--', linewidth=2, label=f'Median: {_money(med)}')
        ax.axvline(mean, color=chart_colors[3], linestyle='--', linewidth=2, label=f'Mean:   {_money(mean)}')
        ax.set_title(
            f'Market value distribution (n={len(prices):,}, top 1% trimmed for readability)',
            fontsize=13,
            fontweight='bold',
            pad=12,
        )
        ax.set_xlabel('Market value')
        ax.set_ylabel('Properties')
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:,.0f}'))
        ax.legend(loc='upper right')
        plt.tight_layout()
        _add_chart(fig)

    # Chart 2: Sqft vs Market Value scatter
    plot_df = df[['building_sqft', 'market_value']].dropna()
    plot_df = plot_df[(plot_df['building_sqft'] > 0) & (plot_df['market_value'] > 0)]
    # cap outliers on both axes
    if len(plot_df) > 5:
        plot_df = plot_df[plot_df['market_value'] <= plot_df['market_value'].quantile(0.99)]
        plot_df = plot_df[plot_df['building_sqft'] <= plot_df['building_sqft'].quantile(0.99)]
        fig, ax = plt.subplots(figsize=(11, 5.5))
        ax.scatter(
            plot_df['building_sqft'], plot_df['market_value'], c=chart_colors[1], alpha=0.4, s=20, edgecolors='none'
        )
        # trend line
        if len(plot_df) > 10:
            z = np.polyfit(plot_df['building_sqft'], plot_df['market_value'], 1)
            xs = np.linspace(plot_df['building_sqft'].min(), plot_df['building_sqft'].max(), 100)
            ax.plot(
                xs,
                np.poly1d(z)(xs),
                color=chart_colors[2],
                linestyle='--',
                linewidth=2,
                label=f'~ ${z[0]:,.0f} per sqft',
            )
            ax.legend()
        ax.set_title(
            f'Building size vs. market value (n={len(plot_df):,}, outliers trimmed)',
            fontsize=13,
            fontweight='bold',
            pad=12,
        )
        ax.set_xlabel('Square footage')
        ax.set_ylabel('Market value')
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:,.0f}'))
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:,.0f}'))
        plt.tight_layout()
        _add_chart(fig)

    # Chart 3: Top property types by count
    type_counts = df['property_type'].value_counts().head(12)
    if len(type_counts):
        # Strip leading code if still present
        labels = [str(t).replace('0', ' ', 0)[:35] for t in type_counts.index]
        fig, ax = plt.subplots(figsize=(11, 5.5))
        bars = ax.barh(
            range(len(type_counts)),
            type_counts.values,
            color=[chart_colors[i % len(chart_colors)] for i in range(len(type_counts))],
        )
        ax.set_yticks(range(len(type_counts)))
        ax.set_yticklabels(labels, fontsize=9)
        ax.invert_yaxis()
        ax.set_title('Top property types by count', fontsize=13, fontweight='bold', pad=12)
        ax.set_xlabel('Number of properties')
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x:,.0f}'))
        for bar, v in zip(bars, type_counts.values, strict=False):
            ax.text(v, bar.get_y() + bar.get_height() / 2, f'  {v:,}', va='center', fontsize=8)
        plt.tight_layout()
        _add_chart(fig)

    # Chart 4: Year built distribution
    years = df['year_built'].dropna()
    years = years[(years > 1900) & (years <= datetime.utcnow().year)]
    if len(years):
        fig, ax = plt.subplots(figsize=(11, 5.5))
        ax.hist(
            years,
            bins=range(int(years.min()), int(years.max()) + 5, 5),
            color=chart_colors[3],
            edgecolor='white',
            alpha=0.85,
        )
        ax.axvline(
            years.median(),
            color=chart_colors[0],
            linestyle='--',
            linewidth=2,
            label=f'Median year: {int(years.median())}',
        )
        ax.set_title('Year built distribution (5-year buckets)', fontsize=13, fontweight='bold', pad=12)
        ax.set_xlabel('Year built')
        ax.set_ylabel('Properties')
        ax.legend()
        plt.tight_layout()
        _add_chart(fig)

    # Chart 5: Average price by city (only useful when not city-filtered)
    show_city_chart = (request is None) or not request.GET.get('city')
    if show_city_chart and 'city' in df.columns:
        city_stats = (
            df.dropna(subset=['city', 'market_value'])
            .groupby('city')['market_value']
            .agg(['mean', 'count'])
            .query('count >= 50')
            .sort_values('mean', ascending=False)
            .head(15)
        )
        if len(city_stats):
            fig, ax = plt.subplots(figsize=(11, 5.5))
            bars = ax.bar(
                range(len(city_stats)),
                city_stats['mean'],
                color=[chart_colors[i % len(chart_colors)] for i in range(len(city_stats))],
                edgecolor='white',
            )
            ax.set_xticks(range(len(city_stats)))
            ax.set_xticklabels(city_stats.index, rotation=35, ha='right', fontsize=9)
            ax.set_title('Mean market value by city (cities with ≥ 50 matches)', fontsize=13, fontweight='bold', pad=12)
            ax.set_ylabel('Mean market value')
            ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:,.0f}'))
            for i, bar in enumerate(bars):
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height(),
                    f'n={int(city_stats["count"].iloc[i]):,}',
                    ha='center',
                    va='bottom',
                    fontsize=8,
                )
            plt.tight_layout()
            _add_chart(fig)

    # Chart 6: Market vs Assessed scatter — useful signal for tax appeal targets
    val_df = df[['market_value', 'assessed_value']].dropna()
    val_df = val_df[(val_df['market_value'] > 0) & (val_df['assessed_value'] > 0)]
    if len(val_df):
        cap = val_df['market_value'].quantile(0.99)
        val_df = val_df[(val_df['market_value'] <= cap) & (val_df['assessed_value'] <= cap)]
        fig, ax = plt.subplots(figsize=(11, 5.5))
        ax.scatter(
            val_df['assessed_value'], val_df['market_value'], c=chart_colors[1], alpha=0.35, s=20, edgecolors='none'
        )
        max_val = max(val_df['market_value'].max(), val_df['assessed_value'].max())
        ax.plot(
            [0, max_val],
            [0, max_val],
            color='grey',
            linestyle='--',
            linewidth=1.5,
            label='Equal value (untaxed gain = 0)',
            alpha=0.7,
        )
        ax.set_title(
            'Market value vs. assessed value (gap = untaxed appreciation)',
            fontsize=13,
            fontweight='bold',
            pad=12,
        )
        ax.set_xlabel('Assessed value')
        ax.set_ylabel('Market value')
        ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:,.0f}'))
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'${x:,.0f}'))
        ax.legend()
        plt.tight_layout()
        _add_chart(fig)

    doc.build(story)
    buf.seek(0)

    filename = f'PinellasMarket_{datetime.utcnow():%Y%m%d}.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# -------- helpers -------------------------------------------------------


def _money(v: float | None) -> str:
    if v is None:
        return '—'
    try:
        return f'${float(v):,.0f}'
    except (TypeError, ValueError):
        return '—'


def _int(v: float | None) -> str:
    if v is None:
        return '—'
    try:
        return f'{int(round(float(v))):,}'
    except (TypeError, ValueError):
        return '—'
