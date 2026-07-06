"""Excel + PDF exports for Pinellas Market Lens."""

from __future__ import annotations

import logging
from datetime import UTC, datetime
from io import BytesIO
from typing import TYPE_CHECKING, Any

from django.contrib import messages
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect
from django.urls import reverse
from openpyxl.styles import Alignment, Font, PatternFill

from .market_insights import build_market_insights
from .palette import BORDER, BORDER_LIGHT, MUTED, PRIMARY, WHITE, openpyxl_rgb

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def generate_excel_response(request: HttpRequest | None = None) -> HttpResponse:
    """Build the analysis workbook, or a friendly error if generation fails."""
    try:
        return _build_excel_response(request)
    except Exception:
        logger.exception('Excel export generation failed')
        return _export_error_response(request, 'Excel workbook')


def _build_excel_response(request: HttpRequest | None) -> HttpResponse:
    import openpyxl
    from openpyxl.utils import get_column_letter

    insights = build_market_insights(request)
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color=openpyxl_rgb(WHITE))
    header_fill = PatternFill('solid', fgColor=openpyxl_rgb(PRIMARY))
    title_font = Font(bold=True, size=14)

    ws = wb.active
    ws.title = 'Overview'
    ws['A1'] = 'Pinellas Market Lens'
    ws['A1'].font = title_font
    ws['A2'] = f'Generated {datetime.now(UTC):%Y-%m-%d %H:%M UTC}'
    ws['A2'].font = Font(italic=True, color=openpyxl_rgb(MUTED))

    row = 4
    row = _write_label_value_section(ws, row, 'Filters', insights['filters'])
    row += 1
    row = _write_label_value_section(ws, row, 'Exact KPIs', [(k['label'], k['value']) for k in insights['kpis']])
    row += 1
    ws.cell(row, 1, 'Analyst Takeaways').font = Font(bold=True)
    row += 1
    for takeaway in insights['takeaways']:
        ws.cell(row, 1, takeaway)
        row += 1
    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 42

    _write_segment_sheet(wb, 'City Segments', insights['city_segments'], header_font, header_fill)
    _write_segment_sheet(wb, 'Property Type Segments', insights['type_segments'], header_font, header_fill)
    _write_outlier_sheet(wb, insights['outliers'], header_font, header_fill)
    _write_sample_sheet(wb, insights['sample_parcels'], header_font, header_fill)
    _write_methodology_sheet(wb, insights['methodology'])

    for sheet in wb.worksheets:
        for column in sheet.columns:
            letter = get_column_letter(column[0].column)
            max_len = max(len(str(cell.value or '')) for cell in column)
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 54)
        for row_cells in sheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'PinellasMarketLens_{datetime.now(UTC):%Y%m%d}.xlsx'
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def generate_pdf_response(request: HttpRequest | None = None) -> HttpResponse:
    """Build the PDF insight brief, or a friendly error if generation fails."""
    try:
        return _build_pdf_response(request)
    except Exception:
        logger.exception('PDF export generation failed')
        return _export_error_response(request, 'PDF brief')


def _build_pdf_response(request: HttpRequest | None) -> HttpResponse:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

    insights = build_market_insights(request)
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=0.55 * inch,
        rightMargin=0.55 * inch,
        topMargin=0.55 * inch,
        bottomMargin=0.55 * inch,
    )

    title_style = ParagraphStyle('Title', fontSize=22, alignment=TA_CENTER, spaceAfter=4)
    subtitle_style = ParagraphStyle('Subtitle', fontSize=10, alignment=TA_CENTER, textColor=rl_colors.grey, spaceAfter=16)
    h_style = ParagraphStyle('Heading', fontSize=13, spaceBefore=12, spaceAfter=6)
    body_style = ParagraphStyle('Body', fontSize=9, alignment=TA_LEFT, spaceAfter=4)

    story: list[Any] = [
        Paragraph('Pinellas Market Lens', title_style),
        Paragraph(
            f'Market insight brief · {insights["exact"]["parcel_count"]:,} filtered parcels · {datetime.now(UTC):%B %d, %Y}',
            subtitle_style,
        ),
    ]

    story.append(Paragraph('Filters', h_style))
    story.append(_pdf_table(insights['filters'], [2.0 * inch, 4.6 * inch]))

    kpi_rows = [(k['label'], k['value']) for k in insights['kpis']]
    story.append(Paragraph('Exact KPIs', h_style))
    story.append(_pdf_table(kpi_rows, [2.5 * inch, 4.1 * inch]))

    story.append(Paragraph('Analyst Takeaways', h_style))
    for takeaway in insights['takeaways']:
        story.append(Paragraph(f'• {takeaway}', body_style))

    story.append(Spacer(1, 0.08 * inch))
    story.append(Paragraph('Top City Segments', h_style))
    story.append(
        _pdf_table(
            [('City', 'Parcels', 'Median Value')] + _segment_pdf_rows(insights['city_segments'][:8]),
            [2.7 * inch, 1.2 * inch, 2.7 * inch],
            header=True,
        )
    )

    story.append(Paragraph('Top Property Type Segments', h_style))
    story.append(
        _pdf_table(
            [('Type', 'Parcels', 'Median Value')] + _segment_pdf_rows(insights['type_segments'][:8]),
            [2.7 * inch, 1.2 * inch, 2.7 * inch],
            header=True,
        )
    )

    story.append(Paragraph('Auditable Outliers', h_style))
    outlier_rows = [('Signal', 'Parcel', 'Metric')]
    for signal, rows in insights['outliers'].items():
        for row in rows[:5]:
            outlier_rows.append(
                (
                    signal.replace('_', ' ').title(),
                    f'{row["address"]}, {row["city"]}',
                    _metric_display(row),
                )
            )
    if len(outlier_rows) == 1:
        outlier_rows.append(('None', 'No outliers for this scope', '-'))
    story.append(_pdf_table(outlier_rows, [1.4 * inch, 3.6 * inch, 1.6 * inch], header=True))

    story.append(Paragraph('Methodology', h_style))
    for note in insights['methodology']:
        story.append(Paragraph(f'• {note}', body_style))

    doc.build(story)
    buf.seek(0)

    filename = f'PinellasMarketLens_{datetime.now(UTC):%Y%m%d}.pdf'
    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _export_error_response(request: HttpRequest | None, kind: str) -> HttpResponse:
    """Fail gracefully: flash a message and return to the dashboard when we can,
    otherwise return a plain friendly error instead of a raw 500 traceback."""
    message = f"We couldn't generate the {kind} right now. Please adjust your filters and try again."
    if request is not None:
        messages.error(request, message)
        return redirect(reverse('insights'))
    return HttpResponse(message, status=500, content_type='text/plain; charset=utf-8')


def _write_label_value_section(ws: Worksheet, start_row: int, title: str, rows: list[tuple[str, str]]) -> int:
    row = start_row
    ws.cell(row, 1, title).font = Font(bold=True)
    row += 1
    for label, value in rows:
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        row += 1
    return row


def _write_segment_sheet(
    wb: Workbook, title: str, rows: list[dict[str, Any]], header_font: Font, header_fill: PatternFill
) -> None:
    ws = wb.create_sheet(title)
    headers = ['Name', 'Count', 'Median Value', 'Mean Value', 'Median $/Sqft', 'Median Tax Rate', 'Mean Assessed Gap %']
    _write_header(ws, headers, header_font, header_fill)
    for idx, row in enumerate(rows, 2):
        ws.cell(idx, 1, row['name'])
        ws.cell(idx, 2, row['count'])
        ws.cell(idx, 3, _number(row['median_value']))
        ws.cell(idx, 4, _number(row['mean_value']))
        ws.cell(idx, 5, _number(row['median_price_per_sqft']))
        ws.cell(idx, 6, _number(row['median_tax_rate']))
        ws.cell(idx, 7, _number(row['mean_assessed_gap_pct']))
    _format_currency_columns(ws, [3, 4, 5])
    _format_percent_columns(ws, [6, 7])


def _write_outlier_sheet(
    wb: Workbook, outliers: dict[str, list[dict[str, Any]]], header_font: Font, header_fill: PatternFill
) -> None:
    ws = wb.create_sheet('Outliers')
    headers = ['Signal', 'Parcel ID', 'Address', 'City', 'Type', 'Market Value', 'Metric Label', 'Metric Value']
    _write_header(ws, headers, header_font, header_fill)
    row_idx = 2
    for signal, rows in outliers.items():
        for row in rows:
            ws.cell(row_idx, 1, signal.replace('_', ' ').title())
            ws.cell(row_idx, 2, row['parcel_id'])
            ws.cell(row_idx, 3, row['address'])
            ws.cell(row_idx, 4, row['city'])
            ws.cell(row_idx, 5, row['property_type'])
            ws.cell(row_idx, 6, _number(row['market_value']))
            ws.cell(row_idx, 7, row['metric_label'])
            ws.cell(row_idx, 8, _number(row['metric_value']))
            row_idx += 1
    _format_currency_columns(ws, [6])


def _write_sample_sheet(
    wb: Workbook, rows: list[dict[str, Any]], header_font: Font, header_fill: PatternFill
) -> None:
    ws = wb.create_sheet('Sample Parcels')
    headers = ['Parcel ID', 'Address', 'City', 'ZIP', 'Type', 'Market Value', 'Assessed Value', '$/Sqft', 'Tax Rate']
    _write_header(ws, headers, header_font, header_fill)
    for idx, row in enumerate(rows, 2):
        ws.cell(idx, 1, row['parcel_id'])
        ws.cell(idx, 2, row['address'])
        ws.cell(idx, 3, row['city'])
        ws.cell(idx, 4, row['zip_code'])
        ws.cell(idx, 5, row['property_type'])
        ws.cell(idx, 6, _number(row['market_value']))
        ws.cell(idx, 7, _number(row['assessed_value']))
        ws.cell(idx, 8, _number(row['price_per_sqft']))
        ws.cell(idx, 9, _number(row['tax_rate']))
    _format_currency_columns(ws, [6, 7, 8])
    _format_percent_columns(ws, [9])


def _write_methodology_sheet(wb: Workbook, notes: list[str]) -> None:
    ws = wb.create_sheet('Methodology')
    ws['A1'] = 'Methodology'
    ws['A1'].font = Font(bold=True, size=14)
    for idx, note in enumerate(notes, 3):
        ws.cell(idx, 1, note)


def _write_header(ws: Worksheet, headers: list[str], header_font: Font, header_fill: PatternFill) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'


def _format_currency_columns(ws: Worksheet, columns: list[int]) -> None:
    for col in columns:
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for item in cell:
                item.number_format = '"$"#,##0'


def _format_percent_columns(ws: Worksheet, columns: list[int]) -> None:
    for col in columns:
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for item in cell:
                item.number_format = '0.00"%"'


def _pdf_table(rows: list[tuple[Any, ...]], col_widths: list[float], header: bool = False):
    from reportlab.lib import colors as rl_colors
    from reportlab.platypus import Table, TableStyle

    table = Table(rows, colWidths=col_widths, repeatRows=1 if header else 0)
    style = [
        ('BOX', (0, 0), (-1, -1), 0.5, rl_colors.HexColor(BORDER)),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, rl_colors.HexColor(BORDER_LIGHT)),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]
    if header:
        style.extend(
            [
                ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor(PRIMARY)),
                ('TEXTCOLOR', (0, 0), (-1, 0), rl_colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ]
        )
    table.setStyle(TableStyle(style))
    return table


def _segment_pdf_rows(rows: list[dict[str, Any]]) -> list[tuple[str, str, str]]:
    if not rows:
        return [('None', '0', '-')]
    return [(row['name'], f'{row["count"]:,}', _money(row['median_value'])) for row in rows]


def _metric_display(row: dict[str, Any]) -> str:
    if row['metric_label'] in ('Tax rate', 'Assessed gap'):
        return _percent(row['metric_value'])
    return _money(row['metric_value'])


def _number(value: Any):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _money(value: Any) -> str:
    number = _number(value)
    if number is None:
        return '-'
    return f'${number:,.0f}'


def _percent(value: Any) -> str:
    number = _number(value)
    if number is None:
        return '-'
    return f'{number:.2f}%'
